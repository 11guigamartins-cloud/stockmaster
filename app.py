from flask import Flask, render_template, request, jsonify, session
from flask_cors import CORS
from datetime import datetime, timedelta
import database as db
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = 'stockmaster_secret_key_2024'
CORS(app)

# Credenciais de admin
ADMIN_USER = 'admin'
ADMIN_PASS = 'belissima1710'

# Inicializar banco de dados
db.init_db()

# ==================== AUTENTICAÇÃO ====================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login do administrador"""
    dados = request.json
    usuario = dados.get('usuario', '')
    senha = dados.get('senha', '')
    
    if usuario == ADMIN_USER and senha == ADMIN_PASS:
        session['autenticado'] = True
        session['timestamp'] = datetime.now().isoformat()
        return jsonify({'sucesso': True, 'mensagem': 'Login realizado com sucesso'}), 200
    else:
        return jsonify({'sucesso': False, 'erro': 'Usuário ou senha inválidos'}), 401

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Logout do administrador"""
    session.clear()
    return jsonify({'sucesso': True, 'mensagem': 'Logout realizado'}), 200

@app.route('/api/auth/verificar', methods=['GET'])
def verificar_autenticacao():
    """Verifica se o usuário está autenticado"""
    autenticado = session.get('autenticado', False)
    return jsonify({'autenticado': autenticado}), 200

def requer_autenticacao(f):
    """Decorator para verificar autenticação"""
    from functools import wraps
    
    @wraps(f)
    def decorador(*args, **kwargs):
        if not session.get('autenticado', False):
            return jsonify({'sucesso': False, 'erro': 'Acesso negado. Faça login primeiro'}), 401
        return f(*args, **kwargs)
    
    return decorador

# ==================== ROTAS DE PRODUTOS ====================

@app.route('/api/produtos', methods=['GET'])
def listar_produtos():
    """Lista produtos com filtros"""
    nome = request.args.get('nome')
    marca = request.args.get('marca')
    linha = request.args.get('linha')
    tipo = request.args.get('tipo')
    codigo = request.args.get('codigo')
    
    produtos = db.listar_produtos(nome, marca, linha, tipo, codigo)
    return jsonify(produtos)

@app.route('/api/produtos', methods=['POST'])
def criar_produto():
    """Cria um novo produto - SEM RESTRIÇÃO"""
    dados = request.json
    
    codigo_limpo = dados['codigobarras'].strip().upper()
    
    id_novo = db.criar_produto(
        codigo_limpo,
        dados['nome'].strip(),
        dados.get('marca', '').strip(),
        dados.get('linha', ''),
        dados.get('tipo', ''),
        float(dados['preco']),
        int(dados['estoque'])
    )
    
    if id_novo:
        return jsonify({'sucesso': True, 'id': id_novo}), 201
    else:
        return jsonify({'sucesso': False, 'erro': 'Código de barras já existe'}), 400

@app.route('/api/produtos/<int:id>', methods=['PUT'])
@requer_autenticacao
def atualizar_produto(id):
    """Atualiza um produto - REQUER LOGIN"""
    dados = request.json
    if db.atualizar_produto(id, **dados):
        return jsonify({'sucesso': True})
    else:
        return jsonify({'sucesso': False}), 400

@app.route('/api/produtos/<int:id>', methods=['DELETE'])
@requer_autenticacao
def deletar_produto(id):
    """Deleta um produto - REQUER LOGIN"""
    if db.deletar_produto(id):
        return jsonify({'sucesso': True})
    else:
        return jsonify({'sucesso': False}), 400
    
@app.route('/api/produtos/editar/<int:id>', methods=['PUT'])
@requer_autenticacao
def editar_produto_completo(id):
    """Edição completa do produto"""
    dados = request.json
    if db.atualizar_produto_completo(id, dados):
        return jsonify({'sucesso': True})
    return jsonify({'sucesso': False, 'erro': 'Erro ao atualizar'}), 400

@app.route('/api/produtos/buscar/<codigo>', methods=['GET'])
def buscar_produto_codigo(codigo):
    """Busca produto pelo código de barras"""
    produto = db.buscar_produto_por_codigo(codigo)
    if produto:
        return jsonify(produto)
    else:
        return jsonify({'erro': 'Produto não encontrado'}), 404

# ==================== ROTAS DE VENDEDORES ====================

@app.route('/api/vendedores', methods=['GET'])
@requer_autenticacao
def listar_vendedores():
    """Lista todos os vendedores"""
    vendedores = db.listar_vendedores()
    return jsonify(vendedores)

@app.route('/api/vendedores', methods=['POST'])
@requer_autenticacao
def criar_vendedor():
    """Cria um novo vendedor"""
    dados = request.json
    
    try:
        id_novo = db.criar_vendedor(
            dados['nome'],
            dados.get('telefone', ''),
            dados['cpf']
        )
        
        if id_novo:
            return jsonify({'sucesso': True, 'id': id_novo}), 201
        else:
            return jsonify({'sucesso': False, 'erro': 'CPF já existe'}), 400
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 400

@app.route('/api/vendedores/<int:id>', methods=['DELETE'])
@requer_autenticacao
def deletar_vendedor(id):
    """Deleta um vendedor"""
    if db.deletar_vendedor(id):
        return jsonify({'sucesso': True})
    else:
        return jsonify({'sucesso': False}), 400

# ==================== ROTAS DE CLIENTES (NOVO) ====================

@app.route('/api/clientes', methods=['GET'])
def listar_clientes():
    """Lista todos os clientes - PUBLICO"""
    clientes = db.listar_clientes()
    return jsonify(clientes)

@app.route('/api/clientes', methods=['POST'])
def criar_cliente():
    """Cria um novo cliente - PUBLICO"""
    dados = request.json
    
    try:
        id_novo = db.criar_cliente(
            dados['nome'],
            dados.get('whatsapp', ''),
            dados['cpf']
        )
        
        if id_novo:
            return jsonify({'sucesso': True, 'id': id_novo}), 201
        else:
            return jsonify({'sucesso': False, 'erro': 'CPF já existe'}), 400
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 400

@app.route('/api/clientes/buscar/<cpf>', methods=['GET'])
def buscar_cliente(cpf):
    """Busca cliente por CPF para o caixa"""
    cliente = db.buscar_cliente_por_cpf(cpf)
    if cliente:
        return jsonify(cliente)
    else:
        return jsonify({'erro': 'Cliente não encontrado'}), 404

@app.route('/api/clientes/<int:id>', methods=['DELETE'])
@requer_autenticacao
def deletar_cliente(id):
    """Deleta um cliente - REQUER LOGIN"""
    if db.deletar_cliente(id):
        return jsonify({'sucesso': True})
    else:
        return jsonify({'sucesso': False}), 400

# ==================== ROTAS DE VENDAS ====================

@app.route('/api/vendas', methods=['GET'])
def listar_vendas():
    """Lista vendas com filtros opcionais"""
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    vendedor_id = request.args.get('vendedor_id')
    produto = request.args.get('produto')
    
    vendas = db.listar_vendas(data_inicio, data_fim, vendedor_id, produto)
    return jsonify(vendas)

@app.route('/api/vendas', methods=['POST'])
def criar_venda():
    """Cria uma nova venda - SEM RESTRIÇÃO"""
    dados = request.json
    
    id_venda = 'VD' + str(int(datetime.now().timestamp()))
    
    sucesso = db.criar_venda(
        id_venda,
        dados['itens'],
        float(dados['subtotal']),
        float(dados.get('descontopercentual', 0)),
        float(dados['total']),
        dados['metodopagamento'],
        dados.get('vendedor_id'),
        dados.get('cliente_id') # NOVO CAMPO
    )
    
    if sucesso:
        venda = db.buscar_venda(id_venda)
        return jsonify({'sucesso': True, 'venda': venda}), 201
    else:
        return jsonify({'sucesso': False, 'erro': 'Erro ao criar venda'}), 400

@app.route('/api/vendas/<venda_id>', methods=['GET'])
def obter_venda(venda_id):
    """Obtém detalhes de uma venda"""
    venda = db.buscar_venda(venda_id)
    if venda:
        return jsonify(venda)
    else:
        return jsonify({'erro': 'Venda não encontrada'}), 404

# ==================== ROTAS DE DASHBOARD ====================

@app.route('/api/dashboard/estatisticas', methods=['GET'])
@requer_autenticacao
def obter_estatisticas():
    """Obtém estatísticas do dashboard - REQUER LOGIN"""
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    vendedor_id = request.args.get('vendedor_id')
    
    stats = db.obter_estatisticas(data_inicio, data_fim, vendedor_id)
    return jsonify(stats)

@app.route('/api/dashboard/vendas-14dias', methods=['GET'])
@requer_autenticacao
def obter_vendas_14dias():
    """Obtém vendas dos últimos 14 dias - REQUER LOGIN"""
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    vendedor_id = request.args.get('vendedor_id')
    
    vendas = db.obter_vendas_14dias(data_inicio, data_fim, vendedor_id)
    return jsonify(vendas)

@app.route('/api/dashboard/top-produtos', methods=['GET'])
@requer_autenticacao
def obter_top_produtos():
    """Obtém top produtos vendidos - REQUER LOGIN"""
    limite = request.args.get('limite', 10, type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    vendedor_id = request.args.get('vendedor_id')
    
    produtos = db.obter_top_produtos(limite, data_inicio, data_fim, vendedor_id)
    return jsonify(produtos)

@app.route('/api/dashboard/ranking-vendedores-valor', methods=['GET'])
@requer_autenticacao
def ranking_vendedores_valor():
    """Ranking de vendedores por valor"""
    limite = request.args.get('limite', 10, type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    ranking = db.obter_ranking_vendedores_valor(limite, data_inicio, data_fim)
    return jsonify(ranking)

@app.route('/api/dashboard/ranking-vendedores-itens', methods=['GET'])
@requer_autenticacao
def ranking_vendedores_itens():
    """Ranking de vendedores por itens"""
    limite = request.args.get('limite', 10, type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    ranking = db.obter_ranking_vendedores_itens(limite, data_inicio, data_fim)
    return jsonify(ranking)

@app.route('/api/dashboard/vendas-por-vendedor-dia', methods=['GET'])
@requer_autenticacao
def vendas_vendedor_dia():
    """Vendas por vendedor por dia"""
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    dados = db.obter_vendas_vendedor_dia(data_inicio, data_fim)
    return jsonify(dados)

# ==================== ROTAS DE RELATÓRIOS ====================

@app.route('/api/relatorio/nota-fiscal/<venda_id>', methods=['GET'])
def gerar_nota_fiscal(venda_id):
    """Gera PDF da nota fiscal"""
    venda = db.buscar_venda(venda_id)
    if not venda:
        return jsonify({'erro': 'Venda não encontrada'}), 404
    
    # Criar PDF em memória
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    width, height = letter
    
    # Cabeçalho
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 50, "BELÍSSIMA COSMÉTICOS")
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 65, "CNPJ: 015.332.279/0001-88")
    c.drawString(50, height - 80, "Salvador, BA")
    
    # Linha
    c.line(50, height - 90, width - 50, height - 90)
    
    # Título do recibo
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, height - 110, "NOTA FISCAL INTERNA")
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 130, f"Pedido: {venda['id']}")
    c.drawString(50, height - 145, f"Data: {venda['data']}")
    
    # Vendedor
    if venda.get('vendedor_nome'):
        c.drawString(50, height - 160, f"Vendedor: {venda['vendedor_nome']}")
        
    # Cliente (NOVO)
    if venda.get('cliente_nome'):
        c.drawString(250, height - 160, f"Cliente: {venda['cliente_nome']}")
    
    # Linha
    c.line(50, height - 175, width - 50, height - 175)
    
    # Cabeçalho da tabela
    y = height - 200
    c.setFont("Helvetica-Bold", 9)
    c.drawString(50, y, "ITEM")
    c.drawString(200, y, "QTD")
    c.drawString(250, y, "VALOR")
    c.drawString(350, y, "TOTAL")
    
    # Itens
    y -= 20
    c.setFont("Helvetica", 9)
    for item in venda['itens']:
        c.drawString(50, y, str(item['nome'])[:30])
        c.drawString(200, y, str(item['quantidade']))
        c.drawString(250, y, f"R$ {float(item['preco']):.2f}")
        c.drawString(350, y, f"R$ {float(item['preco']) * item['quantidade']:.2f}")
        y -= 15
    
    # Linha
    y -= 10
    c.line(50, y, width - 50, y)
    
    # Totais
    y -= 20
    c.setFont("Helvetica", 10)
    c.drawString(250, y, "SUBTOTAL:")
    c.drawString(350, y, f"R$ {float(venda['subtotal']):.2f}")
    
    if float(venda.get('descontopercentual', 0)) > 0:
        y -= 15
        desconto_valor = float(venda['subtotal']) * float(venda['descontopercentual']) / 100
        c.drawString(250, y, f"DESCONTO ({venda['descontopercentual']}%):")
        c.drawString(350, y, f"R$ {desconto_valor:.2f}")
    
    y -= 20
    c.setFont("Helvetica-Bold", 12)
    c.drawString(250, y, "TOTAL:")
    c.drawString(350, y, f"R$ {float(venda['total']):.2f}")
    
    # Método de pagamento
    y -= 40
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y, "PAGAMENTO")
    y -= 15
    c.setFont("Helvetica", 10)
    metodos = {
        'dinheiro': '💵 DINHEIRO',
        'debito': '🏧 DÉBITO',
        'credito': '💳 CRÉDITO',
        'pix': '📱 PIX'
    }
    c.drawString(50, y, metodos.get(venda['metodopagamento'], venda['metodopagamento']))
    
    # Rodapé
    y -= 60
    c.setFont("Helvetica", 9)
    c.drawString(50, y, "Obrigado pela preferência!")
    c.drawString(50, y - 15, "VOLTE SEMPRE.")
    
    c.save()
    
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue(), 200, {
        'Content-Type': 'application/pdf',
        'Content-Disposition': f'attachment; filename=nota_{venda_id}.pdf'
    }

@app.route('/api/relatorio/historico-excel', methods=['GET'])
@requer_autenticacao
def exportar_historico_excel():
    """Exporta histórico em Excel"""
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    vendedor_id = request.args.get('vendedor_id')
    produto = request.args.get('produto')
    
    vendas = db.listar_vendas(data_inicio, data_fim, vendedor_id, produto)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Histórico'
    
    # Headers - ADICIONADO "CLIENTE"
    headers = ['Data/Hora', 'ID Venda', 'Vendedor', 'Cliente', 'Método', 'Produto', 'Qtd', 'Valor Unit', 'Total']
    header_fill = PatternFill(start_color='34495e', end_color='34495e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    row = 2
    for venda in vendas:
        data_fmt = datetime.fromisoformat(venda['data']).strftime('%d/%m/%Y %H:%M') if venda.get('data') else '-'
        
        for item in venda.get('itens', []):
            ws.cell(row=row, column=1).value = data_fmt
            ws.cell(row=row, column=2).value = venda['id']
            ws.cell(row=row, column=3).value = venda.get('vendedor_nome', '-')
            ws.cell(row=row, column=4).value = venda.get('cliente_nome', '-') # NOVO
            ws.cell(row=row, column=5).value = venda['metodopagamento']
            ws.cell(row=row, column=6).value = item['nome']
            ws.cell(row=row, column=7).value = item['quantidade']
            ws.cell(row=row, column=8).value = float(item['preco'])
            ws.cell(row=row, column=9).value = float(item['preco']) * item['quantidade']
            
            row += 1
    
    # Auto adjust columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Salvar
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename=historico_vendas.xlsx'
    }

# ==================== HEALTH CHECK ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Verifica se o servidor está rodando"""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})

# ==================== ROTA RAIZ ====================

@app.route('/')
def index():
    """Serve a página principal"""
    return render_template('index.html')

if __name__ == '__main__':
    print("🚀 StockMaster API com Clientes, Vendedores e Dashboard Avançado rodando em http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)